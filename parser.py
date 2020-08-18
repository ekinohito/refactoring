import pdb
from http.client import HTTPResponse
from os import listdir, curdir
from pickle import dump, load
from typing import List, Dict, Union
from urllib.request import urlopen
from openpyxl import Workbook

# noinspection PyProtectedMember
from bs4 import BeautifulSoup, Tag, NavigableString

Row = Dict[str, Union[int, str, bool]]
Table = List[Row]


def parse_site(file=None):
    url = f'http://admlist.ru/bmstu/5ab880e599b84419489cf11bcf04e3c8.html'
    response: HTTPResponse = urlopen(url)
    print('Got response. Converting...', file=file)

    text = response.read().decode()
    print('Converted response text to string. Processing...', file=file)

    bs = BeautifulSoup(text, features="html.parser")
    print('Made BeautifulSoup. Searching for table...', file=file)

    table: Tag = bs.find_all('table')[-1]
    print('Got table. Parsing table legend...', file=file)

    head: Tag = table.thead
    head_row: Tag = head.tr
    legend: List[Tag] = [cell for cell in head_row.children]
    print('Parsed legend. Parsing table body...', file=file)

    body: Tag = table.tbody
    result: List[Dict[str, Tag]] = list()
    for row in body.children:
        if type(row) == NavigableString:
            continue
        tr: Tag = row
        list_row: Dict[str, Tag] = dict()
        for index, cell in enumerate(tr.children):
            td: Tag = cell
            list_row[legend[index].text] = td
        if len(list_row) == 0:
            continue
        result.append(list_row)
    print('Parsed body. Returning result.', file=file)

    return result


def process_table(raw: List[Dict[str, Tag]]):
    result: Table = list()
    for row in raw:
        result_row: Row = dict()
        result_row['name'] = row['ФИО'].text
        result_row['number'] = int(row['№'].text)
        result_row['accept_here'] = row['Согл'].text == 'Да'
        result_row['type'] = row['Тип'].text
        result_row['sum'] = row['∑'].text
        result_row['accept_else'] = row['Другие ОП'].find('b') is not None
        result.append(result_row)
    return result


def save_to_excel(table: Table, legend: List[str], file_name: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    for horizontal, key in enumerate(legend):
        ws.cell(1, 1 + horizontal, key)
    for vertical, row in enumerate(table):
        for horizontal, key in enumerate(legend):
            ws.cell(2 + vertical, 1 + horizontal, row[key])
    wb.save(file_name)
    wb.close()


def main():
    file_name = 'iu7'
    pickle_name = f'{file_name}.pickle'
    xlsx_name = f'{file_name}.xlsx'
    load_flag: bool = False
    if pickle_name in listdir(curdir):
        load_flag = input('There is a data for this one. Should I use it(y/n)? ').strip() == 'y'

    if load_flag:
        print('Reading data from local copy...')
        with open(pickle_name, 'rb') as fp:
            table = load(fp)
    else:
        print('Reading data from outer source...')
        table = process_table(parse_site())
        with open(pickle_name, 'wb+') as fp:
            dump(table, fp)

    print('Data loaded. Saving to Excel...')
    legend = ['number', 'name', 'type', 'sum', 'accept_here', 'accept_else']
    save_to_excel(table, legend, xlsx_name)

    if input('Data saved. Do you want to continue(y/n)? ') == 'y':
        print('Enjoy!')
        # noinspection PyBroadException
        try:
            pdb.set_trace()
        except Exception as e:
            print(e)
        except BaseException as _:
            print('See you!')

    else:
        print('See you!')


if __name__ == '__main__':
    main()
